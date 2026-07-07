# -*- coding: utf-8 -*-
r"""
把 BigSeller 采集导出的 Excel 转成 TikTok Shop 批量上传模板。

默认规则：
- 商品图 1 沿用 BigSeller 链接。
- 商品图 2-9 替换为 D:\desk\印尼POD\主图 里的 7 张图 + 尺码表。
- 详情 HTML 保留原文字，把原 <img> 替换为 D:\desk\印尼POD\详情 里的全部图片。
- 每个 SKU 库存 999，包裹重量 210g，长宽高 25/22/4cm。
- 正常 SKU 尺码统一补齐为 S/M/L/XL/XXL/XXXL，自动规范 2XL/3XL 等写法。
- 每个商品额外追加一个“不要购买”的加急发货 SKU，默认价格 141300。

如果本地图片已经上传到公网，可以传入 --asset-url-base 或分别传入
--main-url-base / --detail-url-base / --extra-url。

如果要直接上传 Cloudflare R2，传入 --upload-r2，并提供：
--r2-endpoint、--r2-bucket、--public-url-base。访问密钥默认读取环境变量：
CLOUDFLARE_R2_ACCESS_KEY_ID / CLOUDFLARE_R2_SECRET_ACCESS_KEY。
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import hmac
import http.client
import io
import json
import mimetypes
import os
import re
import sys
import time
from collections import OrderedDict
from copy import copy
from pathlib import Path
from typing import Iterable
from urllib.parse import quote, urlparse
from urllib.request import Request, urlopen

import openpyxl


ROOT = Path(__file__).resolve().parent
DEFAULT_MAIN_IMAGE_DIR = Path(r"D:\desk\印尼POD\主图")
DEFAULT_DETAIL_IMAGE_DIR = Path(r"D:\desk\印尼POD\详情")
DEFAULT_EXTRA_IMAGE = Path(
    r"D:\desk\印尼POD\Pengiriman kilat dalam 48 jam (Jangan melakukan pembelian melalui tautan ini).png"
)
EXTRA_SKU_TEXT = "Pengiriman kilat dalam 48 jam (Jangan melakukan pembelian melalui tautan ini)"
EXTRA_VARIATION_VALUE_1 = "Pengiriman kilat 48 jam"
EXTRA_VARIATION_VALUE_2 = "Jangan beli"
DELIVERY_TEXT = "The delivery options for this product are the same as the delivery options for the shop. "
DEFAULT_SIZE_CHART_IMAGE = DEFAULT_DETAIL_IMAGE_DIR / "主8尺码表.png"
MAIN_IMAGE_REPLACEMENT_COUNT = 8
DEFAULT_TARGET_SIZES = ["S", "M", "L", "XL", "XXL", "XXXL"]
SIZE_ALIASES = {
    "S": "S",
    "M": "M",
    "L": "L",
    "XL": "XL",
    "XXL": "XXL",
    "2XL": "XXL",
    "XXXL": "XXXL",
    "3XL": "XXXL",
}
COLOR_CODES = {
    "WHITE": "WH",
    "WH": "WH",
    "PUTIH": "WH",
    "BLACK": "BK",
    "BK": "BK",
    "HITAM": "BK",
    "RED": "RED",
    "MERAH": "RED",
    "BLUE": "BL",
    "BIRU": "BL",
    "GREEN": "GR",
    "HIJAU": "GR",
    "ORANGE": "OG",
    "OREN": "OG",
    "APRICOT": "AP",
    "AP": "AP",
    "GRAY": "GY",
    "GREY": "GY",
    "ABU": "GY",
    "GY": "GY",
}
COLOR_DISPLAY = {
    "WH": "Putih",
    "BK": "Hitam",
    "RED": "Red",
    "BL": "Blue",
    "GR": "Green",
    "OG": "Orange",
    "AP": "Apricot",
    "GY": "Grey",
}
DEFAULT_PRINT_SIDE_OVERRIDE = ROOT / "print_side_overrides.csv"
PRINT_SIDE_PR_KEYWORDS = [
    "front back",
    "front and back",
    "front & back",
    "depan belakang",
    "depan dan belakang",
    "depan & belakang",
    "belakang depan",
    "dua sisi",
    "2 sisi",
    "two side",
    "two sides",
    "both side",
    "both sides",
    "double side",
    "double sided",
    "bolak balik",
    "bolak-balik",
]
PRINT_SIDE_R_KEYWORDS = [
    "back only",
    "hanya belakang",
    "belakang saja",
    "sablon belakang saja",
]
PRINT_SIDE_P_KEYWORDS = [
    "front only",
    "hanya depan",
    "depan saja",
    "satu sisi",
    "1 sisi",
    "one side",
    "single side",
    "single sided",
]

TEMPLATE_SHEET = "Template"
DATA_START_ROW = 6
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}

TEMPLATE_COLUMNS = {
    "category": 1,
    "brand": 2,
    "product_name": 3,
    "product_description": 4,
    "main_image": 5,
    "image_2": 6,
    "image_3": 7,
    "image_4": 8,
    "image_5": 9,
    "image_6": 10,
    "image_7": 11,
    "image_8": 12,
    "image_9": 13,
    "property_name_1": 14,
    "property_value_1": 15,
    "property_1_image": 16,
    "property_name_2": 17,
    "property_value_2": 18,
    "parcel_weight": 19,
    "parcel_length": 20,
    "parcel_width": 21,
    "parcel_height": 22,
    "delivery": 23,
    "price": 24,
    "pre_order_time": 25,
    "quantity": 26,
    "seller_sku": 27,
    "minimum_order_quantity": 28,
    "size_chart": 29,
    "cod": 30,
    "shipping_insurance": 31,
    "materials": 32,
    "pattern": 33,
    "neckline": 34,
    "sleeve_length": 35,
    "season": 36,
    "style": 37,
    "fit": 38,
    "stretch": 39,
    "washing_instructions": 40,
    "waist_height": 41,
}

SOURCE_COLUMNS = {
    "name": "产品名称",
    "long_description": "长描述",
    "short_description": "短描述",
    "source_url": "产品来源链接",
    "variation_name_1": "变种名称 1",
    "variation_value_1": "变种选项 1",
    "variation_name_2": "变种名称 2",
    "variation_value_2": "变种选项 2",
    "price": "价格",
    "sale_price": "促销价",
    "stock": "库存",
    "seller_sku": "SKU",
    "variant_image_1": "变种图 1",
}

PRODUCT_PROPERTIES = {
    "materials": "100%Sorona",
    "pattern": "Graphic",
    "neckline": "Round Neck",
    "sleeve_length": "Short Sleeve",
    "season": "All Seasons",
    "style": "Casual",
    "fit": "Loose-Fitting",
    "stretch": "",
    "washing_instructions": "Machine Washable",
    "waist_height": "",
}

MATERIAL_VALUE = "100%Sorona"
MATERIAL_PATTERNS = [
    (
        re.compile(
            r"\bbahan\s+(?:100\s*%\s*)?(?:cotton|katun)"
            r"(?:\s+(?:combed|premium|pe|20s|24s|30s))*\b",
            re.IGNORECASE,
        ),
        f"Bahan {MATERIAL_VALUE}",
    ),
    (
        re.compile(
            r"\b(?:100\s*%\s*)?(?:cotton|katun)"
            r"(?:\s+(?:combed|premium|pe|20s|24s|30s))*\b",
            re.IGNORECASE,
        ),
        MATERIAL_VALUE,
    ),
    (re.compile(r"\b(?:semi\s+katun|cotton\s+pe|pe\s+cotton)\b", re.IGNORECASE), MATERIAL_VALUE),
]


def natural_key(value: str) -> list[object]:
    stem = Path(value).stem
    suffix = Path(value).suffix.lower()
    parts: list[object] = [
        (1, int(part)) if part.isdigit() else (2, part.lower())
        for part in re.split(r"(\d+)", stem)
        if part
    ]
    return parts + [(0, 0), (3, suffix)]


def find_default_file(pattern: str) -> Path:
    matches = sorted(ROOT.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
    if not matches:
        raise FileNotFoundError(f"找不到匹配文件：{pattern}")
    return matches[0]


def list_images(folder: Path, expected_count: int | None = None) -> list[Path]:
    if not folder.exists():
        raise FileNotFoundError(f"图片文件夹不存在：{folder}")
    images = sorted(
        [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in IMAGE_EXTENSIONS],
        key=lambda p: natural_key(p.name),
    )
    if expected_count is not None and len(images) < expected_count:
        raise ValueError(f"{folder} 至少需要 {expected_count} 张图片，当前只有 {len(images)} 张")
    return images[:expected_count] if expected_count else images


def normalize_local_path(path: Path) -> str:
    return str(path.resolve()).replace("\\", "/")


def join_url(base: str, *parts: str) -> str:
    cleaned = base.rstrip("/")
    encoded_parts = [quote(part.replace("\\", "/").strip("/"), safe="/") for part in parts if part]
    return "/".join([cleaned, *encoded_parts])


def load_asset_manifest(path: Path | None) -> dict[str, object]:
    if not path:
        return {}
    if not path.exists():
        raise FileNotFoundError(f"素材 URL 配置文件不存在：{path}")
    with path.open("r", encoding="utf-8-sig") as handle:
        data = json.load(handle)
    if not isinstance(data, dict):
        raise ValueError("素材 URL 配置文件必须是 JSON object")
    return data


def manifest_string_list(manifest: dict[str, object], key: str) -> list[str]:
    value = manifest.get(key)
    if value is None:
        return []
    if not isinstance(value, list) or not all(isinstance(item, str) and item.strip() for item in value):
        raise ValueError(f"素材 URL 配置里的 {key} 必须是字符串列表")
    return [item.strip() for item in value]


def manifest_urls(manifest: dict[str, object], kind: str, expected_count: int | None = None) -> list[str]:
    direct_urls = manifest_string_list(manifest, f"{kind}_urls")
    if direct_urls:
        urls = direct_urls
    else:
        files = manifest_string_list(manifest, f"{kind}_files")
        if not files:
            urls = []
        else:
            base = as_text(manifest.get(f"{kind}_url_base")) or as_text(manifest.get("asset_url_base"))
            if not base:
                raise ValueError(f"素材 URL 配置使用 {kind}_files 时必须提供 asset_url_base 或 {kind}_url_base")
            urls = [join_url(base, kind, name) for name in files]

    if expected_count is not None and len(urls) < expected_count:
        raise ValueError(f"素材 URL 配置里的 {kind} 至少需要 {expected_count} 个 URL，当前只有 {len(urls)} 个")
    return urls[:expected_count] if expected_count else urls


def manifest_url(manifest: dict[str, object], key: str, kind: str = "") -> str:
    direct_url = as_text(manifest.get(f"{key}_url"))
    if direct_url:
        return direct_url
    filename = as_text(manifest.get(f"{key}_file"))
    if not filename:
        return ""
    base = as_text(manifest.get(f"{kind}_url_base")) if kind else ""
    base = base or as_text(manifest.get("asset_url_base"))
    if not base:
        raise ValueError(f"素材 URL 配置使用 {key}_file 时必须提供 asset_url_base")
    return join_url(base, kind, filename) if kind else join_url(base, filename)


class R2Client:
    def __init__(self, endpoint: str, bucket: str, access_key: str, secret_key: str) -> None:
        self.endpoint = endpoint.rstrip("/")
        self.bucket = bucket
        self.access_key = access_key
        self.secret_key = secret_key
        parsed = urlparse(self.endpoint)
        if parsed.scheme != "https" or not parsed.netloc:
            raise ValueError("--r2-endpoint 必须类似 https://<account_id>.r2.cloudflarestorage.com")
        self.host = parsed.netloc
        self.base_path = parsed.path.rstrip("/")

    def upload_file(self, path: Path, key: str) -> None:
        body = path.read_bytes()
        payload_hash = hashlib.sha256(body).hexdigest()
        now = dt.datetime.now(dt.UTC)
        amz_date = now.strftime("%Y%m%dT%H%M%SZ")
        date_stamp = now.strftime("%Y%m%d")
        content_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        canonical_uri = f"{self.base_path}/{quote(self.bucket, safe='')}/{quote(key, safe='/-_.~')}"
        headers = {
            "host": self.host,
            "x-amz-content-sha256": payload_hash,
            "x-amz-date": amz_date,
            "content-type": content_type,
        }
        signed_headers = ";".join(sorted(headers))
        canonical_headers = "".join(f"{name}:{headers[name]}\n" for name in sorted(headers))
        canonical_request = "\n".join(
            ["PUT", canonical_uri, "", canonical_headers, signed_headers, payload_hash]
        )
        credential_scope = f"{date_stamp}/auto/s3/aws4_request"
        string_to_sign = "\n".join(
            [
                "AWS4-HMAC-SHA256",
                amz_date,
                credential_scope,
                hashlib.sha256(canonical_request.encode("utf-8")).hexdigest(),
            ]
        )
        signing_key = self._signing_key(date_stamp)
        signature = hmac.new(signing_key, string_to_sign.encode("utf-8"), hashlib.sha256).hexdigest()
        headers["authorization"] = (
            "AWS4-HMAC-SHA256 "
            f"Credential={self.access_key}/{credential_scope}, "
            f"SignedHeaders={signed_headers}, Signature={signature}"
        )

        connection = http.client.HTTPSConnection(self.host, timeout=60)
        try:
            connection.request("PUT", canonical_uri, body=body, headers=headers)
            response = connection.getresponse()
            response_body = response.read().decode("utf-8", errors="replace")
            if response.status not in (200, 201):
                raise RuntimeError(f"R2 上传失败 {response.status}: {response_body}")
        finally:
            connection.close()

    def _signing_key(self, date_stamp: str) -> bytes:
        key_date = hmac.new(("AWS4" + self.secret_key).encode("utf-8"), date_stamp.encode("utf-8"), hashlib.sha256).digest()
        key_region = hmac.new(key_date, b"auto", hashlib.sha256).digest()
        key_service = hmac.new(key_region, b"s3", hashlib.sha256).digest()
        return hmac.new(key_service, b"aws4_request", hashlib.sha256).digest()


class AssetResolver:
    def __init__(self, args: argparse.Namespace) -> None:
        self.args = args
        self.r2_client: R2Client | None = None
        if args.upload_r2:
            access_key = args.r2_access_key or os.getenv("CLOUDFLARE_R2_ACCESS_KEY_ID")
            secret_key = args.r2_secret_key or os.getenv("CLOUDFLARE_R2_SECRET_ACCESS_KEY")
            missing = [
                name
                for name, value in {
                    "--r2-endpoint": args.r2_endpoint,
                    "--r2-bucket": args.r2_bucket,
                    "--public-url-base": args.public_url_base,
                    "CLOUDFLARE_R2_ACCESS_KEY_ID": access_key,
                    "CLOUDFLARE_R2_SECRET_ACCESS_KEY": secret_key,
                }.items()
                if not value
            ]
            if missing:
                raise ValueError("开启 --upload-r2 时缺少配置：" + ", ".join(missing))
            self.r2_client = R2Client(args.r2_endpoint, args.r2_bucket, access_key, secret_key)

    def resolve_many(self, kind: str, paths: Iterable[Path]) -> list[str]:
        return [self.resolve(kind, path) for path in paths]

    def resolve(self, kind: str, path: Path) -> str:
        if self.r2_client:
            key = self._r2_key(kind, path)
            self.r2_client.upload_file(path, key)
            return join_url(self.args.public_url_base, key)

        if kind == "main" and self.args.main_url_base:
            return join_url(self.args.main_url_base, path.name)
        if kind == "detail" and self.args.detail_url_base:
            return join_url(self.args.detail_url_base, path.name)
        if kind == "extra" and self.args.extra_url:
            return self.args.extra_url
        if self.args.asset_url_base:
            return join_url(self.args.asset_url_base, kind, path.name)
        return normalize_local_path(path)

    def _r2_key(self, kind: str, path: Path) -> str:
        prefix = self.args.r2_prefix.strip("/")
        key = f"{kind}/{path.name}"
        return f"{prefix}/{key}" if prefix else key


def index_headers(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if value:
            headers[str(value).strip()] = col
    return headers


def cell_value(row: dict[str, object], header: str) -> object:
    return row.get(header)


def as_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def as_number(value: object, fallback: int | float = 0) -> int | float:
    if value is None or value == "":
        return fallback
    if isinstance(value, (int, float)):
        return value
    text = str(value).replace(",", "").strip()
    try:
        number = float(text)
    except ValueError:
        return fallback
    return int(number) if number.is_integer() else number


def normalize_size(value: object) -> str:
    text = as_text(value).upper()
    text = re.sub(r"[\s_\-]+", "", text)
    return SIZE_ALIASES.get(text, "")


def parse_target_sizes(value: str) -> list[str]:
    sizes = [normalize_size(part) for part in value.split(",")]
    sizes = [size for size in sizes if size]
    if not sizes:
        raise ValueError("尺码列表不能为空，示例：S,M,L,XL,XXL,XXXL")
    result = []
    for size in sizes:
        if size not in result:
            result.append(size)
    return result


def parse_color_codes(value: str) -> list[str]:
    codes = []
    for part in value.split(","):
        code = normalize_color_code(part) or as_text(part).upper()
        if code and code not in codes:
            codes.append(code)
    return codes


def normalize_color_code(value: object) -> str:
    text = as_text(value).upper()
    if not text:
        return ""
    compact = re.sub(r"[^0-9A-Z]+", " ", text)
    tokens = [token for token in compact.split() if token]
    for token in tokens:
        if token in COLOR_CODES:
            return COLOR_CODES[token]
    for token, code in COLOR_CODES.items():
        if token and token in compact:
            return code
    return ""


def infer_color_code(row: dict[str, object], default_color_code: str) -> str:
    for key in ("variation_value_1", "variation_value_2", "name"):
        code = normalize_color_code(row.get(SOURCE_COLUMNS[key]))
        if code:
            return code
    return normalize_color_code(default_color_code) or default_color_code.upper()


def normalize_print_side(value: str, default: str = "P") -> str:
    side = re.sub(r"[^A-Z]+", "", as_text(value).upper())
    if side == "AUTO":
        return "AUTO"
    if side == "RP":
        return "PR"
    return side if side in {"P", "R", "PR"} else default


def normalize_keyword_text(value: object) -> str:
    text = as_text(value).lower()
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"[^0-9a-z]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def find_keyword(text: str, keywords: list[str]) -> str:
    padded = f" {text} "
    for keyword in keywords:
        normalized = normalize_keyword_text(keyword)
        if normalized and f" {normalized} " in padded:
            return keyword
    return ""


def load_print_side_overrides(path: Path) -> dict[str, str]:
    if not path or not path.exists():
        return {}
    overrides: dict[str, str] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            side = normalize_print_side(row.get("print_side", ""), default="")
            if side not in {"P", "R", "PR"}:
                continue
            for key_name in ("product_no", "product_index", "source_url", "product_name", "title"):
                key_value = as_text(row.get(key_name))
                if key_value:
                    overrides[f"{key_name}:{normalize_keyword_text(key_value)}"] = side
    return overrides


def product_text(rows: list[dict[str, object]]) -> str:
    parts = []
    for row in rows:
        for source_key in ("name", "seller_sku", "short_description", "long_description"):
            parts.append(as_text(row.get(SOURCE_COLUMNS[source_key])))
    return normalize_keyword_text(" ".join(parts))


def print_side_image_candidates(rows: list[dict[str, object]], max_images: int) -> list[str]:
    urls = []
    seen = set()

    def add(value: object) -> None:
        url = as_text(value)
        if not url.lower().startswith(("http://", "https://")) or url in seen:
            return
        seen.add(url)
        urls.append(url)

    for row in rows:
        add(row.get(SOURCE_COLUMNS["variant_image_1"]))
        if len(urls) >= max_images:
            return urls
    for row in rows:
        for url in source_images(row):
            add(url)
            if len(urls) >= max_images:
                return urls
    return urls


def download_image(url: str, timeout: int) -> bytes:
    last_error: Exception | None = None
    for attempt in range(1, 4):
        request = Request(
            url,
            headers={
                "User-Agent": "Mozilla/5.0",
                "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
                "Connection": "close",
            },
        )
        try:
            with urlopen(request, timeout=timeout) as response:
                content_type = response.headers.get("Content-Type", "")
                if "image" not in content_type.lower():
                    return b""
                return response.read(8 * 1024 * 1024)
        except Exception as exc:
            last_error = exc
            if attempt < 3:
                time.sleep(0.4 * attempt)
    if last_error:
        raise last_error
    return b""


def connected_components(mask) -> list[dict[str, int]]:
    import numpy as np

    h, w = mask.shape
    visited = np.zeros(mask.shape, dtype=bool)
    components = []
    for y in range(h):
        for x in range(w):
            if not mask[y, x] or visited[y, x]:
                continue
            stack = [(y, x)]
            visited[y, x] = True
            area = 0
            min_x = max_x = x
            min_y = max_y = y
            while stack:
                cy, cx = stack.pop()
                area += 1
                min_x = min(min_x, cx)
                max_x = max(max_x, cx)
                min_y = min(min_y, cy)
                max_y = max(max_y, cy)
                for ny, nx in ((cy - 1, cx), (cy + 1, cx), (cy, cx - 1), (cy, cx + 1)):
                    if 0 <= ny < h and 0 <= nx < w and mask[ny, nx] and not visited[ny, nx]:
                        visited[ny, nx] = True
                        stack.append((ny, nx))
            components.append(
                {
                    "area": area,
                    "min_x": min_x,
                    "max_x": max_x,
                    "min_y": min_y,
                    "max_y": max_y,
                    "cx": (min_x + max_x) // 2,
                    "cy": (min_y + max_y) // 2,
                }
            )
    return components


def image_has_side_by_side_garments(image_bytes: bytes) -> tuple[bool, str]:
    try:
        import numpy as np
        from PIL import Image, ImageFilter
    except Exception:
        return False, "PIL/numpy unavailable"

    try:
        image = Image.open(io.BytesIO(image_bytes)).convert("RGB")
    except Exception as exc:
        return False, f"cannot open image: {exc}"

    image.thumbnail((240, 240))
    arr = np.asarray(image).astype("int16")
    h, w, _ = arr.shape
    if h < 80 or w < 80:
        return False, "image too small"

    border = np.concatenate([arr[:5, :, :].reshape(-1, 3), arr[-5:, :, :].reshape(-1, 3), arr[:, :5, :].reshape(-1, 3), arr[:, -5:, :].reshape(-1, 3)])
    bg = np.median(border, axis=0)
    distance = np.sqrt(((arr - bg) ** 2).sum(axis=2))
    contrast = arr.max(axis=2) - arr.min(axis=2)
    mask = (distance > 34) | ((contrast > 38) & (distance > 18))
    mask[:3, :] = False
    mask[-3:, :] = False
    mask[:, :3] = False
    mask[:, -3:] = False

    mask_img = Image.fromarray((mask.astype("uint8") * 255), mode="L")
    mask_img = mask_img.filter(ImageFilter.MedianFilter(3)).filter(ImageFilter.MaxFilter(3)).filter(ImageFilter.MinFilter(3))
    mask = np.asarray(mask_img) > 0
    total = int(mask.sum())
    if total < h * w * 0.025:
        return False, "foreground too small"

    components = connected_components(mask)
    large = []
    for component in components:
        width = component["max_x"] - component["min_x"] + 1
        height = component["max_y"] - component["min_y"] + 1
        if component["area"] >= h * w * 0.018 and width >= w * 0.08 and height >= h * 0.24:
            large.append(component)
    large.sort(key=lambda item: item["area"], reverse=True)

    mid_x = w // 2
    mid_y = h // 2
    quadrants = {
        "top-left": int(mask[:mid_y, :mid_x].sum()),
        "top-right": int(mask[:mid_y, mid_x:].sum()),
        "bottom-left": int(mask[mid_y:, :mid_x].sum()),
        "bottom-right": int(mask[mid_y:, mid_x:].sum()),
    }
    diagonal_pairs = [
        ("top-left/bottom-right", quadrants["top-left"], quadrants["bottom-right"]),
        ("top-right/bottom-left", quadrants["top-right"], quadrants["bottom-left"]),
    ]
    for label, first_mass, second_mass in diagonal_pairs:
        diagonal_ratio = (first_mass + second_mass) / total
        other_ratio = 1 - diagonal_ratio
        smaller_ratio = min(first_mass, second_mass) / total
        if diagonal_ratio >= 0.72 and smaller_ratio >= 0.28 and diagonal_ratio - other_ratio >= 0.34:
            return True, (
                f"two diagonal foreground masses; orientation={label}, "
                f"diagonal_ratio={diagonal_ratio:.2f}, smaller_ratio={smaller_ratio:.2f}"
            )

    for i, left in enumerate(large[:5]):
        for right in large[i + 1 : 5]:
            separation = abs(left["cx"] - right["cx"]) / w
            overlap_y = min(left["max_y"], right["max_y"]) - max(left["min_y"], right["min_y"])
            overlap_ratio = overlap_y / max(1, min(left["max_y"] - left["min_y"] + 1, right["max_y"] - right["min_y"] + 1))
            area_ratio = min(left["area"], right["area"]) / max(1, max(left["area"], right["area"]))
            if separation >= 0.22 and overlap_ratio >= 0.25 and area_ratio >= 0.28:
                return True, f"two large separated foreground objects; separation={separation:.2f}, area_ratio={area_ratio:.2f}"

    x_projection = mask.sum(axis=0).astype("float32")
    if x_projection.max() <= 0:
        return False, "no foreground projection"
    window = max(5, w // 28)
    kernel = np.ones(window, dtype="float32") / window
    smooth = np.convolve(x_projection, kernel, mode="same")
    left_peak = smooth[: w // 2].max()
    right_peak = smooth[w // 2 :].max()
    middle = smooth[int(w * 0.42) : int(w * 0.58)].mean()
    left_mass = mask[:, : w // 2].sum()
    right_mass = mask[:, w // 2 :].sum()
    if (
        left_mass >= total * 0.24
        and right_mass >= total * 0.24
        and min(left_peak, right_peak) > 0
        and middle <= min(left_peak, right_peak) * 0.66
    ):
        return True, f"two horizontal foreground masses; middle_valley={middle / min(left_peak, right_peak):.2f}"

    return False, "single dominant foreground or no clear side-by-side split"


def detect_print_side_from_images(rows: list[dict[str, object]], args: argparse.Namespace) -> dict[str, str] | None:
    if not args.image_print_side_detect:
        return None
    for url in print_side_image_candidates(rows, args.image_detect_max_images):
        if url not in args.image_analysis_cache:
            try:
                image_bytes = download_image(url, args.image_detect_timeout)
                if image_bytes:
                    args.image_analysis_cache[url] = image_has_side_by_side_garments(image_bytes)
                else:
                    args.image_analysis_cache[url] = (False, "downloaded content is not an image")
            except Exception as exc:
                args.image_analysis_cache[url] = (False, f"download failed: {exc}")
        matched, reason = args.image_analysis_cache[url]
        if matched:
            return {
                "side": "PR",
                "confidence": "medium",
                "method": "image",
                "evidence": f"{reason}; {url}",
            }
    return None


def image_base_color_code(image_bytes: bytes) -> tuple[str, str]:
    try:
        import numpy as np
        from PIL import Image
    except Exception:
        return "", "PIL/numpy unavailable"
    try:
        image = Image.open(io.BytesIO(image_bytes)).convert("RGB")
    except Exception as exc:
        return "", f"cannot open image: {exc}"
    image.thumbnail((240, 240))
    arr = np.asarray(image).astype("int16")
    h, w, _ = arr.shape
    if h < 80 or w < 80:
        return "", "image too small"
    crop = arr[int(h * 0.18) : int(h * 0.92), int(w * 0.05) : int(w * 0.95), :]
    luminance = 0.299 * crop[:, :, 0] + 0.587 * crop[:, :, 1] + 0.114 * crop[:, :, 2]
    dark_ratio = float((luminance < 130).mean())
    very_dark_ratio = float((luminance < 80).mean())
    median_luminance = float(np.median(luminance))
    if dark_ratio >= 0.22 or very_dark_ratio >= 0.12 or median_luminance < 190:
        return "BK", f"dark_ratio={dark_ratio:.2f}, very_dark_ratio={very_dark_ratio:.2f}, median_luminance={median_luminance:.0f}"
    return "WH", f"dark_ratio={dark_ratio:.2f}, very_dark_ratio={very_dark_ratio:.2f}, median_luminance={median_luminance:.0f}"


def detect_color_code_from_images(rows: list[dict[str, object]], args: argparse.Namespace) -> tuple[str, str]:
    if not args.image_color_detect:
        return "", "disabled"
    for url in print_side_image_candidates(rows, args.image_detect_max_images):
        if url not in args.image_color_cache:
            try:
                image_bytes = download_image(url, args.image_detect_timeout)
                if image_bytes:
                    args.image_color_cache[url] = image_base_color_code(image_bytes)
                else:
                    args.image_color_cache[url] = ("", "downloaded content is not an image")
            except Exception as exc:
                args.image_color_cache[url] = ("", f"download failed: {exc}")
        code, reason = args.image_color_cache[url]
        if code in {"BK", "WH"}:
            return code, f"{reason}; {url}"
    return "", "no usable image color signal"


def infer_group_color_code(rows: list[dict[str, object]], args: argparse.Namespace) -> tuple[str, str]:
    for row in rows:
        for key in ("variation_value_1", "variation_value_2", "name"):
            code = normalize_color_code(row.get(SOURCE_COLUMNS[key]))
            if code:
                return code, f"text:{key}"
    code, reason = detect_color_code_from_images(rows, args)
    if code:
        return code, f"image:{reason}"
    fallback = normalize_color_code(args.default_color_code) or args.default_color_code.upper()
    return fallback, "default"


def detect_print_side(
    rows: list[dict[str, object]],
    args: argparse.Namespace,
    group_index: int,
) -> dict[str, str]:
    forced_side = normalize_print_side(args.print_side, default="AUTO")
    if forced_side != "AUTO":
        return {
            "side": forced_side,
            "confidence": "forced",
            "method": "argument",
            "evidence": f"--print-side {forced_side}",
        }

    first_row = rows[0]
    override_candidates = [
        f"product_no:{group_index:03d}",
        f"product_index:{group_index:03d}",
        f"product_index:{group_index}",
        f"source_url:{normalize_keyword_text(first_row.get(SOURCE_COLUMNS['source_url']))}",
        f"product_name:{normalize_keyword_text(first_row.get(SOURCE_COLUMNS['name']))}",
        f"title:{normalize_keyword_text(first_row.get(SOURCE_COLUMNS['name']))}",
    ]
    for key in override_candidates:
        if key in args.print_side_overrides:
            side = args.print_side_overrides[key]
            return {
                "side": side,
                "confidence": "override",
                "method": "override",
                "evidence": key,
            }

    text = product_text(rows)
    for side, keywords in (
        ("R", PRINT_SIDE_R_KEYWORDS),
        ("PR", PRINT_SIDE_PR_KEYWORDS),
        ("P", PRINT_SIDE_P_KEYWORDS),
    ):
        matched = find_keyword(text, keywords)
        if matched:
            return {
                "side": side,
                "confidence": "high",
                "method": "keyword",
                "evidence": matched,
            }

    image_result = detect_print_side_from_images(rows, args)
    if image_result:
        return image_result

    return {
        "side": "P",
        "confidence": "low",
        "method": "fallback",
        "evidence": "no print-side keyword matched",
    }


def write_print_side_review(path: Path, rows: list[dict[str, object]]) -> None:
    if not rows:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["product_no", "print_side", "confidence", "method", "evidence", "product_name", "source_url"]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def sku_prefix(value: str) -> str:
    prefix = re.sub(r"[^0-9A-Z]+", "", as_text(value).upper())
    return prefix[:8] or "X"


def sku_date(value: str) -> str:
    text = re.sub(r"[^0-9]+", "", as_text(value))
    if len(text) == 8:
        return text
    return dt.datetime.now().strftime("%Y%m%d")


def sku_date_from_source_filename(path: Path) -> str:
    for match in re.finditer(r"(20\d{6})", path.stem):
        value = match.group(1)
        try:
            dt.datetime.strptime(value, "%Y%m%d")
        except ValueError:
            continue
        return value
    return ""


def resolve_sku_date(value: str, source_path: Path) -> str:
    explicit = re.sub(r"[^0-9]+", "", as_text(value))
    if len(explicit) == 8:
        return explicit
    return sku_date_from_source_filename(source_path) or sku_date(value)


def read_source_rows(path: Path) -> list[dict[str, object]]:
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb.active
    headers = index_headers(ws)
    missing = [name for name in SOURCE_COLUMNS.values() if name not in headers and name != "SKU"]
    if missing:
        raise ValueError("BigSeller 表缺少必要列：" + ", ".join(missing))

    rows = []
    for row_index in range(2, ws.max_row + 1):
        item = {
            header: ws.cell(row_index, col).value
            for header, col in headers.items()
        }
        if as_text(item.get(SOURCE_COLUMNS["name"])):
            item["_row_index"] = row_index
            rows.append(item)
    if not rows:
        raise ValueError("BigSeller 表里没有可转换的数据行")
    return rows


def group_rows(rows: list[dict[str, object]]) -> OrderedDict[str, list[dict[str, object]]]:
    grouped: OrderedDict[str, list[dict[str, object]]] = OrderedDict()
    for row in rows:
        source_url = as_text(row.get(SOURCE_COLUMNS["source_url"]))
        product_name = as_text(row.get(SOURCE_COLUMNS["name"]))
        key = source_url or product_name
        grouped.setdefault(key, []).append(row)
    return grouped


def replace_description_images(description: str, detail_urls: list[str]) -> str:
    detail_tags = "\n".join(f'<img src="{url}">' for url in detail_urls)
    if not description:
        return detail_tags

    image_pattern = re.compile(r"<img\b[^>]*>", flags=re.IGNORECASE)
    replaced = False

    def replace_once(_: re.Match[str]) -> str:
        nonlocal replaced
        if replaced:
            return ""
        replaced = True
        return detail_tags

    if image_pattern.search(description):
        return image_pattern.sub(replace_once, description)

    body_match = re.search(r"<body[^>]*>", description, flags=re.IGNORECASE)
    if body_match:
        insert_at = body_match.end()
        return description[:insert_at] + "\n" + detail_tags + description[insert_at:]
    return detail_tags + "\n" + description


def truncate(text: str, limit: int) -> str:
    return text if len(text) <= limit else text[:limit]


def normalize_material_text(text: str) -> str:
    if not text:
        return text
    normalized = text
    for pattern, replacement in MATERIAL_PATTERNS:
        normalized = pattern.sub(replacement, normalized)
    return re.sub(r"[ \t]{2,}", " ", normalized)


def clean_sku_part(text: str) -> str:
    text = re.sub(r"\s+", "-", text.strip())
    text = re.sub(r"[^0-9A-Za-z_-]+", "", text)
    return text[:30] or "SKU"


def normal_sku(row: dict[str, object], args: argparse.Namespace, group_index: int) -> str:
    size = as_text(row.get("_sku_size")) or normalize_size(row.get(SOURCE_COLUMNS["variation_value_1"])) or normalize_size(row.get(SOURCE_COLUMNS["variation_value_2"]))
    color = as_text(row.get("_sku_color_code")) or infer_color_code(row, args.default_color_code)
    side = normalize_print_side(as_text(row.get("_print_side")) or args.print_side, default="P")
    sku = f"{sku_prefix(args.sku_prefix)}-{sku_date(args.sku_date)}{group_index:03d}-{side}-{color}-{size}"
    return sku[:50]


def extra_sku(args: argparse.Namespace, group_index: int) -> str:
    sku = f"{sku_prefix(args.sku_prefix)}-{sku_date(args.sku_date)}{group_index:03d}-NB"
    return sku[:50]


def source_images(row: dict[str, object]) -> list[str]:
    return [as_text(row.get(f"产品图 {index}")) for index in range(1, 10)]


def detect_size_key(rows: list[dict[str, object]]) -> str:
    candidates = ["variation_value_1", "variation_value_2"]
    scores: dict[str, int] = {}
    for key in candidates:
        header = SOURCE_COLUMNS[key]
        scores[key] = sum(1 for row in rows if normalize_size(row.get(header)))
    best_key = max(candidates, key=lambda key: (scores[key], key == "variation_value_2"))
    return best_key if scores[best_key] else ""


def paired_variation_key(value_key: str, target: str) -> str:
    number = value_key.rsplit("_", 1)[-1]
    return f"variation_{target}_{number}"


def other_variation_value_key(value_key: str) -> str:
    return "variation_value_2" if value_key == "variation_value_1" else "variation_value_1"


def ordered_unique_values(rows: list[dict[str, object]], source_key: str) -> list[str]:
    seen = set()
    values = []
    header = SOURCE_COLUMNS[source_key]
    for row in rows:
        value = as_text(row.get(header))
        if value and value not in seen:
            seen.add(value)
            values.append(value)
    return values


def expand_rows_to_target_sizes(
    rows: list[dict[str, object]],
    target_sizes: list[str],
    args: argparse.Namespace,
) -> list[dict[str, object]]:
    def output_color_groups() -> list[dict[str, object]]:
        allowed_colors = parse_color_codes(args.allowed_colors)
        selected = [
            code
            for code in parse_color_codes(args.output_colors)
            if code in allowed_colors
        ]
        if not selected:
            raise ValueError("--output-colors 至少要包含 BK 或 WH 中的一个")
        return [
            {"code": code, "label": COLOR_DISPLAY.get(code, code)}
            for code in selected
        ]

    selected_groups = output_color_groups()
    source_image_color = normalize_color_code(args.source_image_color) or args.source_image_color.upper()
    if source_image_color not in parse_color_codes(args.allowed_colors):
        source_image_color = "BK"

    size_key = detect_size_key(rows)
    if not size_key:
        expanded = []
        for color_group in selected_groups:
            color_code = as_text(color_group["code"])
            for size in target_sizes:
                new_row = dict(rows[0])
                if len(selected_groups) > 1:
                    new_row[SOURCE_COLUMNS["variation_name_1"]] = "Warna"
                    new_row[SOURCE_COLUMNS["variation_value_1"]] = color_group["label"]
                    new_row[SOURCE_COLUMNS["variation_name_2"]] = "Ukuran"
                    new_row[SOURCE_COLUMNS["variation_value_2"]] = size
                else:
                    new_row[SOURCE_COLUMNS["variation_name_1"]] = "Ukuran"
                    new_row[SOURCE_COLUMNS["variation_value_1"]] = size
                    new_row[SOURCE_COLUMNS["variation_name_2"]] = None
                    new_row[SOURCE_COLUMNS["variation_value_2"]] = None
                new_row["_sku_color_code"] = color_code
                new_row["_sku_size"] = size
                new_row["_sku_variant_image_allowed"] = color_code == source_image_color
                expanded.append(new_row)
        return expanded

    size_header = SOURCE_COLUMNS[size_key]
    size_name_header = SOURCE_COLUMNS[paired_variation_key(size_key, "name")]
    other_key = other_variation_value_key(size_key)
    other_header = SOURCE_COLUMNS[other_key]
    other_name_header = SOURCE_COLUMNS[paired_variation_key(other_key, "name")]

    source_color_codes: list[str] = []
    seen_colors = set()
    for row in rows:
        color_code = normalize_color_code(row.get(other_header))
        if color_code not in parse_color_codes(args.allowed_colors) or color_code in seen_colors:
            continue
        seen_colors.add(color_code)
        source_color_codes.append(color_code)
    other_is_color = bool(source_color_codes)

    expanded = []
    for color_group in selected_groups:
        color_code = as_text(color_group["code"])
        has_source_color = other_is_color and color_code in source_color_codes
        candidates = [row for row in rows if normalize_color_code(row.get(other_header)) == color_code] if has_source_color else rows
        valid_candidates = [row for row in candidates if normalize_size(row.get(size_header))]
        template_fallback = valid_candidates[0] if valid_candidates else (candidates[0] if candidates else rows[0])

        for size in target_sizes:
            template = next(
                (row for row in valid_candidates if normalize_size(row.get(size_header)) == size),
                template_fallback,
            )
            new_row = dict(template)
            new_row[size_header] = size
            new_row[size_name_header] = as_text(new_row.get(size_name_header)) or "Ukuran"
            new_row["_sku_color_code"] = color_code
            new_row["_sku_size"] = size
            new_row["_sku_variant_image_allowed"] = has_source_color or (not other_is_color and color_code == source_image_color)
            if other_is_color or len(selected_groups) > 1:
                new_row[other_header] = as_text(color_group["label"]) or COLOR_DISPLAY.get(color_code, color_code)
                new_row[other_name_header] = "Warna"
            else:
                new_row[other_header] = None
                new_row[other_name_header] = None
            expanded.append(new_row)
    return expanded


def row_price(row: dict[str, object]) -> int | float:
    sale_price = as_number(row.get(SOURCE_COLUMNS["sale_price"]))
    return sale_price or as_number(row.get(SOURCE_COLUMNS["price"]))


def base_template_row(
    row: dict[str, object],
    args: argparse.Namespace,
    product_images: list[str],
    description: str,
    group_index: int,
    row_number: int,
) -> dict[str, object]:
    variation_name_1 = as_text(row.get(SOURCE_COLUMNS["variation_name_1"]))
    variation_name_2 = as_text(row.get(SOURCE_COLUMNS["variation_name_2"]))
    variant_image = ""
    if row.get("_sku_variant_image_allowed") is not False:
        variant_image = as_text(row.get(SOURCE_COLUMNS["variant_image_1"]))

    output = {
        "category": args.category,
        "brand": args.brand,
        "product_name": truncate(normalize_material_text(as_text(row.get(SOURCE_COLUMNS["name"]))), 254),
        "product_description": normalize_material_text(description),
        "main_image": product_images[0],
        "image_2": product_images[1],
        "property_name_1": variation_name_1,
        "property_value_1": truncate(as_text(row.get(SOURCE_COLUMNS["variation_value_1"])), 50),
        "property_1_image": variant_image,
        "property_name_2": variation_name_2,
        "property_value_2": truncate(as_text(row.get(SOURCE_COLUMNS["variation_value_2"])), 50),
        "parcel_weight": args.weight,
        "parcel_length": args.length,
        "parcel_width": args.width,
        "parcel_height": args.height,
        "delivery": DELIVERY_TEXT,
        "price": args.normal_price if args.normal_price is not None else row_price(row),
        "quantity": args.stock,
        "seller_sku": normal_sku(row, args, group_index),
        "size_chart": args.size_chart_resolved_url,
        "shipping_insurance": args.shipping_insurance,
    }
    output.update(PRODUCT_PROPERTIES)
    return output


def extra_sku_row(
    first_row: dict[str, object],
    args: argparse.Namespace,
    product_images: list[str],
    description: str,
    extra_image_url: str,
    group_rows_for_price: list[dict[str, object]],
    group_index: int,
    multi_product: bool,
) -> dict[str, object]:
    prices = [row_price(row) for row in group_rows_for_price if row_price(row)]
    base_price = args.normal_price if args.normal_price is not None else (min(prices) if prices else 0)
    price = args.extra_price if args.extra_price is not None else max(base_price - args.extra_price_discount, 0)
    variation_name_1 = as_text(first_row.get(SOURCE_COLUMNS["variation_name_1"]))
    variation_name_2 = as_text(first_row.get(SOURCE_COLUMNS["variation_name_2"]))

    output = {
        "category": args.category,
        "brand": args.brand,
        "product_name": truncate(normalize_material_text(as_text(first_row.get(SOURCE_COLUMNS["name"]))), 254),
        "product_description": normalize_material_text(description),
        "main_image": extra_image_url,
        "image_2": product_images[1],
        "property_name_1": variation_name_1,
        "property_value_1": EXTRA_VARIATION_VALUE_1 if variation_name_1 else "",
        "property_1_image": extra_image_url,
        "property_name_2": variation_name_2,
        "property_value_2": EXTRA_VARIATION_VALUE_2 if variation_name_2 else "",
        "parcel_weight": args.weight,
        "parcel_length": args.length,
        "parcel_width": args.width,
        "parcel_height": args.height,
        "delivery": DELIVERY_TEXT,
        "price": price,
        "quantity": args.stock,
        "seller_sku": extra_sku(args, group_index),
        "size_chart": args.size_chart_resolved_url,
        "shipping_insurance": args.shipping_insurance,
    }
    output.update(PRODUCT_PROPERTIES)
    return output


def write_template_row(ws, row_index: int, data: dict[str, object]) -> None:
    for key, col in TEMPLATE_COLUMNS.items():
        ws.cell(row_index, col).value = data.get(key, "")


def clear_data_area(ws, start_row: int) -> None:
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).value = None


def copy_row_format(ws, source_row: int, target_row: int) -> None:
    if target_row == source_row:
        return
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(source_row, col)
        target_cell = ws.cell(target_row, col)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)


def build_output_rows(
    source_rows: list[dict[str, object]],
    args: argparse.Namespace,
    main_urls: list[str],
    detail_urls: list[str],
    extra_image_url: str,
) -> list[dict[str, object]]:
    grouped = group_rows(source_rows)
    output_rows: list[dict[str, object]] = []
    target_sizes = parse_target_sizes(args.target_sizes)
    multi_product = len(grouped) > 1
    for group_index, rows in enumerate(grouped.values(), start=1):
        first_row = rows[0]
        print_side_info = detect_print_side(rows, args, group_index)
        args.print_side_review_rows.append(
            {
                "product_no": f"{group_index:03d}",
                "print_side": print_side_info["side"],
                "confidence": print_side_info["confidence"],
                "method": print_side_info["method"],
                "evidence": print_side_info["evidence"],
                "product_name": as_text(first_row.get(SOURCE_COLUMNS["name"])),
                "source_url": as_text(first_row.get(SOURCE_COLUMNS["source_url"])),
            }
        )
        original_images = source_images(first_row)
        if not original_images[0]:
            raise ValueError(f"源表第 {first_row['_row_index']} 行缺少产品图 1")

        product_images = [original_images[0], *main_urls]
        description = replace_description_images(
            as_text(first_row.get(SOURCE_COLUMNS["long_description"]))
            or as_text(first_row.get(SOURCE_COLUMNS["short_description"])),
            detail_urls,
        )

        sku_rows = expand_rows_to_target_sizes(rows, target_sizes, args)
        for row_number, row in enumerate(sku_rows, start=1):
            row["_print_side"] = print_side_info["side"]
            output_rows.append(
                base_template_row(row, args, product_images, description, group_index, row_number)
            )
            for image_index, image_url in enumerate(main_urls[1:], start=3):
                output_rows[-1][f"image_{image_index}"] = image_url

        extra = extra_sku_row(first_row, args, product_images, description, extra_image_url, sku_rows, group_index, multi_product)
        for image_index, image_url in enumerate(main_urls[1:], start=3):
            extra[f"image_{image_index}"] = image_url
        output_rows.append(extra)
    return output_rows


def convert(args: argparse.Namespace) -> Path:
    template_path = args.template or find_default_file("*_template.xlsx")
    source_path = args.source or find_default_file("scraped_product_*.xlsx")
    args.sku_date = resolve_sku_date(args.sku_date, source_path)
    output_path = args.output
    if output_path is None:
        timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = ROOT / "outputs" / f"tiktokshop_import_{timestamp}.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    args.print_side_overrides = load_print_side_overrides(args.print_side_override)
    args.print_side_review_rows = []
    args.image_analysis_cache = {}
    args.image_color_cache = {}
    args.print_side_review_path = args.print_side_review or output_path.with_name(
        f"{output_path.stem}_print_side_review.csv"
    )

    asset_manifest = load_asset_manifest(args.asset_manifest)
    if asset_manifest:
        main_urls = manifest_urls(asset_manifest, "main", expected_count=MAIN_IMAGE_REPLACEMENT_COUNT)
        detail_urls = manifest_urls(asset_manifest, "detail")
        extra_image_url = manifest_url(asset_manifest, "extra", "extra")
        args.size_chart_resolved_url = args.size_chart_url or manifest_url(asset_manifest, "size_chart", "detail")
        if not detail_urls:
            raise ValueError("素材 URL 配置里的 detail_urls/detail_files 不能为空")
        if not extra_image_url:
            raise ValueError("素材 URL 配置必须提供 extra_url 或 extra_file")
        if not args.size_chart_resolved_url:
            raise ValueError("素材 URL 配置必须提供 size_chart_url 或 size_chart_file")
    else:
        main_images = list_images(args.main_image_dir, expected_count=7)
        detail_images = list_images(args.detail_image_dir)
        if not detail_images:
            raise ValueError(f"详情图片文件夹没有图片：{args.detail_image_dir}")
        if not args.extra_sku_image.exists():
            raise FileNotFoundError(f"加急 SKU 图片不存在：{args.extra_sku_image}")

        if not args.size_chart_url and not args.size_chart_image.exists():
            raise FileNotFoundError(f"Size chart image not found: {args.size_chart_image}")

        resolver = AssetResolver(args)
        main_urls = resolver.resolve_many("main", main_images)
        detail_urls = resolver.resolve_many("detail", detail_images)
        extra_image_url = resolver.resolve("extra", args.extra_sku_image)
        args.size_chart_resolved_url = args.size_chart_url or resolver.resolve("detail", args.size_chart_image)
        main_urls = [*main_urls, args.size_chart_resolved_url]

    source_rows = read_source_rows(source_path)
    output_rows = build_output_rows(source_rows, args, main_urls, detail_urls, extra_image_url)

    wb = openpyxl.load_workbook(template_path)
    if TEMPLATE_SHEET not in wb.sheetnames:
        raise ValueError(f"模板里找不到工作表：{TEMPLATE_SHEET}")
    ws = wb[TEMPLATE_SHEET]

    clear_data_area(ws, args.start_row)
    for offset, data in enumerate(output_rows):
        row_index = args.start_row + offset
        if row_index > args.start_row:
            copy_row_format(ws, args.start_row, row_index)
        write_template_row(ws, row_index, data)

    wb.save(output_path)
    write_print_side_review(args.print_side_review_path, args.print_side_review_rows)
    print(f"源 SKU 行数：{len(source_rows)}")
    print(f"商品链接数：{len(group_rows(source_rows))}")
    print(f"输出 SKU 行数：{len(output_rows)}")
    print(f"主图替换数：{len(main_urls)}")
    print(f"详情图替换数：{len(detail_urls)}")
    print(f"SKU 日期：{args.sku_date}")
    if not (args.asset_manifest or args.asset_url_base or args.main_url_base or args.detail_url_base or args.extra_url or args.upload_r2):
        print("提示：当前本地素材写入为本地路径。正式上传 TikTok 前，请使用公网 URL 或开启 R2 上传。")
    print(f"已生成：{output_path}")
    print(f"印花面审核清单：{args.print_side_review_path}")
    return output_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="BigSeller Excel 转 TikTok Shop 批量上传模板")
    parser.add_argument("--template", type=Path, default=None, help="TikTok Shop 批量上传模板 xlsx")
    parser.add_argument("--source", type=Path, default=None, help="BigSeller 采集导出 xlsx")
    parser.add_argument("--output", type=Path, default=None, help="输出 xlsx 路径")
    parser.add_argument("--main-image-dir", type=Path, default=DEFAULT_MAIN_IMAGE_DIR)
    parser.add_argument("--detail-image-dir", type=Path, default=DEFAULT_DETAIL_IMAGE_DIR)
    parser.add_argument("--extra-sku-image", type=Path, default=DEFAULT_EXTRA_IMAGE)
    parser.add_argument("--size-chart-image", type=Path, default=DEFAULT_SIZE_CHART_IMAGE)
    parser.add_argument("--size-chart-url", default="")
    parser.add_argument("--category", default="Men's Tops/T-shirts")
    parser.add_argument("--brand", default="No brand")
    parser.add_argument("--stock", type=int, default=999)
    parser.add_argument("--weight", type=int, default=210)
    parser.add_argument("--length", type=int, default=25)
    parser.add_argument("--width", type=int, default=22)
    parser.add_argument("--height", type=int, default=4)
    parser.add_argument("--normal-price", type=int, default=259000, help="正常衣服 SKU 的统一售价；传空值不可用，默认 259000")
    parser.add_argument("--extra-price", type=int, default=141300, help="不要购买 SKU 的固定售价，默认 141300")
    parser.add_argument("--extra-price-discount", type=int, default=3)
    parser.add_argument("--target-sizes", default="S,M,L,XL,XXL,XXXL", help="最终输出尺码，默认 S,M,L,XL,XXL,XXXL")
    parser.add_argument("--sku-prefix", default="X")
    parser.add_argument("--sku-date", default="")
    parser.add_argument("--print-side", default="auto")
    parser.add_argument("--print-side-override", type=Path, default=DEFAULT_PRINT_SIDE_OVERRIDE)
    parser.add_argument("--print-side-review", type=Path, default=None)
    parser.add_argument("--image-print-side-detect", dest="image_print_side_detect", action="store_true", default=True)
    parser.add_argument("--no-image-print-side-detect", dest="image_print_side_detect", action="store_false")
    parser.add_argument("--image-color-detect", dest="image_color_detect", action="store_true", default=True)
    parser.add_argument("--no-image-color-detect", dest="image_color_detect", action="store_false")
    parser.add_argument("--image-detect-max-images", type=int, default=4)
    parser.add_argument("--image-detect-timeout", type=int, default=12)
    parser.add_argument("--allowed-colors", default="BK,WH")
    parser.add_argument("--output-colors", default="BK,WH", help="最终生成的底色 SKU，默认 BK,WH")
    parser.add_argument("--source-image-color", default="BK", help="源表没有明确颜色列时，源 SKU 图属于哪个底色，默认 BK")
    parser.add_argument("--default-color-code", default="BK")
    parser.add_argument("--shipping-insurance", default="Optional")
    parser.add_argument("--start-row", type=int, default=DATA_START_ROW)

    parser.add_argument("--asset-url-base", default="", help="公网素材根地址，会拼成 main/detail/extra/文件名")
    parser.add_argument("--main-url-base", default="", help="主图文件夹对应的公网地址")
    parser.add_argument("--detail-url-base", default="", help="详情图文件夹对应的公网地址")
    parser.add_argument("--extra-url", default="", help="加急 SKU 图片的公网 URL")
    parser.add_argument("--asset-manifest", type=Path, default=None, help="GitHub Actions 等无本地素材环境使用的素材 URL JSON")

    parser.add_argument("--upload-r2", action="store_true", help="上传本地素材到 Cloudflare R2")
    parser.add_argument("--r2-endpoint", default="", help="R2 S3 endpoint，例如 https://<account_id>.r2.cloudflarestorage.com")
    parser.add_argument("--r2-bucket", default="", help="R2 bucket 名称")
    parser.add_argument("--r2-prefix", default="indonesia-pod", help="R2 对象前缀")
    parser.add_argument("--public-url-base", default="", help="R2 公开访问根地址")
    parser.add_argument("--r2-access-key", default="", help="默认读取 CLOUDFLARE_R2_ACCESS_KEY_ID")
    parser.add_argument("--r2-secret-key", default="", help="默认读取 CLOUDFLARE_R2_SECRET_ACCESS_KEY")
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    try:
        convert(args)
    except Exception as exc:
        print(f"转换失败：{exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
